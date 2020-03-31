# import openpyxl module 
import openpyxl 

import sys
import re
import os.path
from os import path
  
# import BarChart class from openpyxl.chart sub_module 
from openpyxl.chart import LineChart,Reference 

def Read_Two_Column_File(file_name):
    with open(file_name, 'r') as data:
        x = []
        y = []
        for line in data:
            p = line.split()
            x.append(p[0])
            y.append(p[1])

    return x, y

x, y = Read_Two_Column_File(sys.argv[1])

# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet  
# from the active attribute. 
ws = wb.active 

# write o to 9 in 1st column of the active sheet 

ws.cell(1,1).value = "CPU Util"
ws.cell(1,2).value = "CPU WUtil"
for i, cpuval in enumerate(x):
    ws.cell(row=i+2, column=1).value = int(cpuval)

for i, cpuval in enumerate(y):
    ws.cell(row=i+2, column=2).value = int(cpuval)
  
c1 = LineChart()
c1.title = "SPC3 SPU Core Load %"
 
c1.y_axis.title = 'CPU %'
c1.x_axis.title = 'CPU Cores'

data = Reference(ws, min_col=1, min_row=1, max_col=2, max_row=40)
c1.add_data(data, titles_from_data=True)
 

ws.add_chart(c1, "D2")

filename = sys.argv[1] + "-chart.xlsx"
wb.save(filename)

